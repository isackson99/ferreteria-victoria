from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q
from decimal import Decimal, InvalidOperation
import io
import csv
from .models import Producto, Categoria, Kit, KitComponente, MovimientoInventario
from .serializers import ProductoSerializer, CategoriaSerializer, KitComponenteSerializer
from usuarios.permissions import PuedeVerInventario, PuedeCrearProductos


# ── Helpers para importación ──────────────────────────────────────────────────

_ALIAS = {
    'codigo':         ['codigo', 'code', 'barcode', 'cod'],
    'nombre':         ['nombre', 'descripcion', 'description', 'name', 'producto'],
    'tipo':           ['tipo', 'type'],
    'precio_costo':   ['precio_costo', 'costo', 'cost', 'preciocosto'],
    'precio_venta':   ['precio_venta', 'venta', 'precio', 'price', 'precioventa'],
    'precio_mayoreo': ['precio_mayoreo', 'mayoreo', 'preciomayoreo'],
    'minimo_mayoreo': ['minimo_mayoreo', 'mayoreo_minimo', 'minimomayoreo', 'minmayoreo'],
    'categoria':      ['categoria', 'departamento', 'category', 'dept', 'department'],
    'stock':          ['stock', 'inventario', 'existencia', 'stock_inicial', 'cantidad'],
    'stock_minimo':   ['stock_minimo', 'inventario_minimo', 'stockminimo', 'minstock'],
    'stock_maximo':   ['stock_maximo', 'inventario_maximo', 'stockmaximo', 'maxstock'],
}


def _norm(s):
    return str(s or '').lower().replace(' ', '').replace('_', '')


def _parse_archivo(archivo):
    """Lee xlsx o csv y retorna (headers, rows_iter)."""
    nombre = archivo.name.lower()
    if nombre.endswith('.xlsx') or nombre.endswith('.xls'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError('openpyxl no está instalado. Ejecuta: pip install openpyxl')
        wb = load_workbook(io.BytesIO(archivo.read()), data_only=True)
        ws = wb.active
        rows = list(ws.values)
        if not rows:
            return [], []
        headers = [str(h or '').strip() for h in rows[0]]
        return headers, rows[1:]
    elif nombre.endswith('.csv'):
        content = archivo.read().decode('utf-8-sig')
        reader = list(csv.reader(io.StringIO(content)))
        if not reader:
            return [], []
        return [str(h).strip() for h in reader[0]], reader[1:]
    else:
        raise ValueError('Formato no soportado. Use .xlsx o .csv')


def _map_columns(headers):
    """Retorna dict: field_name → column_index."""
    norm_headers = [_norm(h) for h in headers]
    mapping = {}
    for field, aliases in _ALIAS.items():
        for alias in aliases:
            if _norm(alias) in norm_headers:
                mapping[field] = norm_headers.index(_norm(alias))
                break
    return mapping


def _safe_decimal(val, default=None):
    if val is None or str(val).strip() == '':
        return default
    try:
        return Decimal(str(val).strip().replace(',', '.'))
    except (InvalidOperation, ValueError):
        return default


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset = Categoria.objects.filter(activa=True)
    serializer_class = CategoriaSerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [PuedeCrearProductos()]
        return [PuedeVerInventario()]

    @action(detail=False, methods=['get'])
    def todas(self, request):
        """Incluye categorías inactivas — solo Admin"""
        categorias = Categoria.objects.all()
        return Response(CategoriaSerializer(categorias, many=True).data)


class ProductoViewSet(viewsets.ModelViewSet):
    serializer_class = ProductoSerializer
    filterset_fields = ['categoria', 'tipo', 'codigo']

    def get_queryset(self):
        qs = Producto.objects.filter(activo=True).select_related('categoria')
        query = self.request.query_params.get('search', '').strip()
        if query:
            if '%' in query:
                # Wildcard mode: remove % and search anywhere (icontains)
                q = query.replace('%', '').strip()
                if q:
                    qs = qs.filter(Q(nombre__icontains=q) | Q(codigo__icontains=q))
            else:
                # Default: search from beginning (istartswith)
                qs = qs.filter(Q(nombre__istartswith=query) | Q(codigo__istartswith=query))
        return qs

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy',
                           'importar', 'preview_importacion']:
            return [PuedeCrearProductos()]
        return [PuedeVerInventario()]

    def _sync_componentes(self, producto, componentes_data):
        """Replace all kit components for a producto of tipo 'kit'."""
        kit, _ = Kit.objects.get_or_create(producto=producto)
        kit.componentes.all().delete()
        for comp_data in componentes_data:
            try:
                comp_id = comp_data.get('componente') or comp_data.get('componente_id')
                componente = Producto.objects.get(id=comp_id, activo=True)
                if componente.id == producto.id:
                    continue  # no self-reference
                KitComponente.objects.create(
                    kit=kit,
                    componente=componente,
                    cantidad=comp_data.get('cantidad', 1)
                )
            except Producto.DoesNotExist:
                pass

    def _refresh_producto(self, producto_id):
        return Producto.objects.select_related('categoria').prefetch_related(
            'kit__componentes__componente'
        ).get(id=producto_id)

    def create(self, request, *args, **kwargs):
        componentes_data = request.data.get('componentes', None)
        response = super().create(request, *args, **kwargs)
        if response.status_code == 201 and componentes_data is not None:
            producto = Producto.objects.get(id=response.data['id'])
            if producto.tipo == 'kit':
                self._sync_componentes(producto, componentes_data)
                producto = self._refresh_producto(producto.id)
                response.data = self.get_serializer(producto).data
        return response

    def update(self, request, *args, **kwargs):
        componentes_data = request.data.get('componentes', None)
        response = super().update(request, *args, **kwargs)
        if componentes_data is not None:
            producto = Producto.objects.get(id=response.data['id'])
            if producto.tipo == 'kit':
                self._sync_componentes(producto, componentes_data)
                producto = self._refresh_producto(producto.id)
                response.data = self.get_serializer(producto).data
        return response

    def destroy(self, request, *args, **kwargs):
        from django.db.models import ProtectedError
        from ventas.models import DetalleVenta

        producto = self.get_object()
        referencias = DetalleVenta.objects.filter(producto=producto).count()

        if referencias > 0:
            producto.activo = False
            producto.save()
            return Response({
                'mensaje': (
                    f'El producto "{producto.nombre}" fue desactivado porque '
                    f'tiene {referencias} referencia(s) en ventas históricas. '
                    f'No aparecerá en búsquedas ni en ventas nuevas.'
                ),
                'desactivado': True
            }, status=200)

        producto.delete()
        return Response({
            'mensaje': f'Producto "{producto.nombre}" eliminado correctamente.',
            'eliminado': True
        }, status=200)

    @action(detail=False, methods=['get'])
    def sin_costo(self, request):
        """Productos activos con precio_costo = 0."""
        productos = self.get_queryset().filter(precio_costo=0)
        return Response({'count': productos.count()})

    @action(detail=True, methods=['get'])
    def componentes_kit(self, request, pk=None):
        producto = self.get_object()
        if producto.tipo != 'kit' or not hasattr(producto, 'kit'):
            return Response([])
        componentes = producto.kit.componentes.select_related('componente').all()
        return Response(KitComponenteSerializer(componentes, many=True).data)

    @action(detail=False, methods=['get'])
    def buscar(self, request):
        query = request.query_params.get('q', '')
        productos = self.get_queryset().filter(nombre__icontains=query) | \
                    self.get_queryset().filter(codigo__icontains=query)
        return Response(self.get_serializer(productos, many=True).data)

    @action(detail=True, methods=['patch'])
    def ajustar_inventario(self, request, pk=None):
        from usuarios.permissions import PuedeAjustarInventario
        from productos.models import MovimientoInventario
        from decimal import Decimal, InvalidOperation

        if not PuedeAjustarInventario().has_permission(request, self):
            return Response({'error': 'No tienes permiso para ajustar inventario.'}, status=403)

        producto = self.get_object()

        if producto.tipo == 'kit':
            return Response(
                {'error': 'Los kits no tienen inventario propio. Ajuste el stock de sus componentes.'},
                status=400
            )

        # Accept nueva_cantidad (preferred) or inventario_actual (legacy)
        raw = request.data.get('nueva_cantidad') or request.data.get('inventario_actual')
        if raw is None:
            return Response({'error': 'Debes enviar nueva_cantidad.'}, status=400)

        try:
            nueva_cantidad = Decimal(str(raw))
        except (InvalidOperation, TypeError):
            return Response({'error': 'Valor de cantidad inválido.'}, status=400)

        if nueva_cantidad < 0:
            return Response({'error': 'La cantidad no puede ser negativa.'}, status=400)

        tipo = request.data.get('tipo', 'ajuste')
        if tipo not in ('entrada', 'salida', 'ajuste'):
            tipo = 'ajuste'

        motivo = request.data.get('motivo', '') or ('Entrada de mercadería' if tipo == 'entrada' else 'Ajuste manual')

        stock_antes = producto.inventario_actual
        producto.inventario_actual = nueva_cantidad

        # Update prices if provided and different
        for campo in ('precio_venta', 'precio_mayoreo'):
            val = request.data.get(campo)
            if val is not None:
                try:
                    dec = Decimal(str(val))
                    if campo == 'precio_venta' and dec > 0:
                        producto.precio_venta = dec
                    elif campo == 'precio_mayoreo':
                        producto.precio_mayoreo = dec if dec > 0 else None
                except (InvalidOperation, TypeError):
                    pass

        producto.save()

        MovimientoInventario.objects.create(
            producto=producto,
            tipo=tipo,
            cantidad=abs(nueva_cantidad - stock_antes),
            stock_antes=stock_antes,
            stock_despues=nueva_cantidad,
            motivo=motivo,
            usuario=request.user,
        )

        from productos.signals import notificar_stock_actualizado, crear_notificacion_stock
        notificar_stock_actualizado(producto)
        crear_notificacion_stock(producto)

        return Response(self.get_serializer(producto).data)

    # ── Importación masiva ────────────────────────────────────────────────────

    @action(detail=False, methods=['get'])
    def plantilla_importacion(self, request):
        """Genera y descarga una plantilla Excel para importación masiva."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            return Response({'error': 'openpyxl no instalado. Ejecuta: pip install openpyxl'}, status=500)

        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = 'Productos'

        headers = [
            'Codigo', 'Nombre', 'Tipo', 'PrecioCosto', 'PrecioVenta',
            'PrecioMayoreo', 'MinimoMayoreo', 'Categoria', 'Stock',
            'StockMinimo', 'StockMaximo',
        ]
        hdr_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
        hdr_font = Font(color='FFFFFF', bold=True, size=11)

        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')

        example = ['TORN001', 'Tornillo 1/2" x 1"', 'unidad', 25.00, 50.00, 45.00, 12,
                   'Ferretería General', 100, 10, 500]
        for i, v in enumerate(example, 1):
            ws.cell(row=2, column=i, value=v)

        for i, w in enumerate([12, 32, 10, 12, 12, 14, 14, 22, 10, 12, 12], 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        # Hoja instrucciones
        ws2 = wb.create_sheet('Instrucciones')
        filas_inst = [
            ('Columna', 'Descripción', 'Obligatorio', 'Valores válidos'),
            ('Codigo', 'Código único del producto', 'SÍ', 'Texto, máx 50 caracteres'),
            ('Nombre', 'Descripción del producto', 'SÍ', 'Texto, máx 200 caracteres'),
            ('Tipo', 'Tipo de venta', 'NO', 'unidad / granel / kit  (default: unidad)'),
            ('PrecioCosto', 'Precio de costo', 'NO', 'Número decimal (ej: 25.50)'),
            ('PrecioVenta', 'Precio de venta al público', 'SÍ', 'Número decimal mayor a 0'),
            ('PrecioMayoreo', 'Precio especial por cantidad', 'NO', 'Número decimal'),
            ('MinimoMayoreo', 'Cantidad mínima para precio mayoreo', 'NO', 'Número entero'),
            ('Categoria', 'Departamento / Categoría', 'NO', 'Texto — se creará si no existe'),
            ('Stock', 'Stock inicial', 'NO', 'Número (default: 0)'),
            ('StockMinimo', 'Stock mínimo (alerta)', 'NO', 'Número (default: 0)'),
            ('StockMaximo', 'Stock máximo', 'NO', 'Número (default: 0)'),
        ]
        hdr2_fill = PatternFill(start_color='37474F', end_color='37474F', fill_type='solid')
        for i, row in enumerate(filas_inst, 1):
            for j, val in enumerate(row, 1):
                cell = ws2.cell(row=i, column=j, value=val)
                if i == 1:
                    cell.fill = hdr2_fill
                    cell.font = Font(color='FFFFFF', bold=True)
                    cell.alignment = Alignment(horizontal='center')
        for col, w in zip(['A', 'B', 'C', 'D'], [16, 45, 12, 42]):
            ws2.column_dimensions[col].width = w

        for r, txt in enumerate([
            'NOTAS IMPORTANTES:',
            '• El Codigo es el identificador único. Si ya existe, el producto se ACTUALIZARÁ.',
            '• Si la Categoria no existe, se creará automáticamente.',
            '• Si Stock > 0, se registrará una entrada de inventario.',
            '• Los encabezados son flexibles: Nombre/Descripcion, Categoria/Departamento, Stock/Inventario',
        ], start=14):
            ws2.cell(row=r, column=1, value=txt)
        ws2['A14'].font = Font(bold=True, color='C62828')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_productos.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def preview_importacion(self, request):
        """Retorna las primeras 5 filas del archivo para previsualización."""
        if 'archivo' not in request.FILES:
            return Response({'error': 'No se envió ningún archivo.'}, status=400)
        try:
            headers, rows = _parse_archivo(request.FILES['archivo'])
        except ValueError as e:
            return Response({'error': str(e)}, status=400)
        except Exception as e:
            return Response({'error': f'Error al leer el archivo: {e}'}, status=400)

        col_map = _map_columns(headers)
        preview = [[str(v) if v is not None else '' for v in row] for row in list(rows)[:5]]
        return Response({
            'columnas': headers,
            'filas': preview,
            'columnas_reconocidas': list(col_map.keys()),
        })

    @action(detail=False, methods=['post'])
    def importar(self, request):
        """Importa productos desde Excel o CSV (upsert por código)."""
        if 'archivo' not in request.FILES:
            return Response({'error': 'No se envió ningún archivo.'}, status=400)
        try:
            headers, rows = _parse_archivo(request.FILES['archivo'])
        except ValueError as e:
            return Response({'error': str(e)}, status=400)
        except Exception as e:
            return Response({'error': f'Error al leer el archivo: {e}'}, status=400)

        col_map = _map_columns(headers)
        requeridos = [f for f in ('codigo', 'nombre', 'precio_venta') if f not in col_map]
        if requeridos:
            return Response({'error': f'Columnas requeridas no encontradas: {", ".join(requeridos)}'}, status=400)

        creados = actualizados = 0
        errores = []

        for fila_num, row in enumerate(rows, start=2):
            row = list(row)

            def get_val(field, default=None):
                idx = col_map.get(field)
                if idx is None or idx >= len(row):
                    return default
                v = row[idx]
                return default if (v is None or str(v).strip() == '') else v

            try:
                codigo = str(get_val('codigo', '')).strip()
                nombre = str(get_val('nombre', '')).strip()
                if not codigo:
                    errores.append({'fila': fila_num, 'motivo': 'Código vacío'}); continue
                if not nombre:
                    errores.append({'fila': fila_num, 'motivo': 'Nombre vacío'}); continue

                tipo_raw = str(get_val('tipo', 'unidad')).strip().lower()
                tipo = tipo_raw if tipo_raw in ('unidad', 'granel', 'kit') else 'unidad'

                precio_venta = _safe_decimal(get_val('precio_venta'))
                if precio_venta is None or precio_venta <= 0:
                    errores.append({'fila': fila_num, 'motivo': 'PrecioVenta inválido o <= 0'}); continue

                precio_costo   = _safe_decimal(get_val('precio_costo'), Decimal('0'))
                precio_mayoreo = _safe_decimal(get_val('precio_mayoreo'))
                min_mayoreo    = _safe_decimal(get_val('minimo_mayoreo'))
                stock          = max(Decimal('0'), _safe_decimal(get_val('stock'), Decimal('0')))
                stock_min      = max(Decimal('0'), _safe_decimal(get_val('stock_minimo'), Decimal('0')))
                stock_max      = max(Decimal('0'), _safe_decimal(get_val('stock_maximo'), Decimal('0')))

                # Categoria
                categoria = None
                cat_nombre = get_val('categoria')
                if cat_nombre and str(cat_nombre).strip():
                    cat_nombre = str(cat_nombre).strip()
                    qs_cat = Categoria.objects.filter(nombre__iexact=cat_nombre)
                    categoria = qs_cat.first() if qs_cat.exists() else Categoria.objects.create(nombre=cat_nombre)

                # Upsert
                existing = Producto.objects.filter(codigo=codigo).first()
                if existing:
                    stock_antes = existing.inventario_actual
                    existing.nombre        = nombre
                    existing.tipo          = tipo
                    existing.precio_venta  = precio_venta
                    existing.precio_costo  = precio_costo
                    existing.precio_mayoreo   = precio_mayoreo
                    existing.mayoreo_minimo   = min_mayoreo
                    existing.inventario_minimo = stock_min
                    existing.inventario_maximo = stock_max
                    existing.activo = True
                    if categoria is not None:
                        existing.categoria = categoria
                    if stock > 0:
                        existing.inventario_actual = stock
                    existing.save()
                    if stock > 0:
                        MovimientoInventario.objects.create(
                            producto=existing, tipo='entrada', cantidad=stock,
                            stock_antes=stock_antes, stock_despues=stock,
                            motivo='Importación masiva', usuario=request.user,
                        )
                    actualizados += 1
                else:
                    nuevo = Producto.objects.create(
                        codigo=codigo, nombre=nombre, tipo=tipo,
                        precio_venta=precio_venta, precio_costo=precio_costo,
                        precio_mayoreo=precio_mayoreo, mayoreo_minimo=min_mayoreo,
                        categoria=categoria, inventario_actual=stock,
                        inventario_minimo=stock_min, inventario_maximo=stock_max,
                        usa_inventario=(tipo != 'kit'),
                    )
                    if stock > 0:
                        MovimientoInventario.objects.create(
                            producto=nuevo, tipo='entrada', cantidad=stock,
                            stock_antes=Decimal('0'), stock_despues=stock,
                            motivo='Importación masiva', usuario=request.user,
                        )
                    creados += 1

            except Exception as e:
                errores.append({'fila': fila_num, 'motivo': str(e)})

        return Response({'creados': creados, 'actualizados': actualizados, 'errores': errores})